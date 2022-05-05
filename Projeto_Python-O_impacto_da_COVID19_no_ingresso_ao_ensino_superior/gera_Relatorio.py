import os
import wget
import math
import openpyxl
import pandas as pd
import matplotlib.pyplot as plt
import requests
from bs4 import BeautifulSoup
from fpdf import FPDF

# Links dos dados dos Enem
urlInsc2019 = requests.get('https://www.gov.br/inep/pt-br/assuntos/noticias/enem/mais-de-5-milhoes-de-participantes-estao-confirmados-para-a-edicao-do-exame-em-2019')
contentInsc2019 = urlInsc2019.content

siteInsc2019 = BeautifulSoup(contentInsc2019, 'html.parser')
tbodyInsc2019 = siteInsc2019.find('tbody')

urlAus2019 = requests.get('https://www.gov.br/inep/pt-br/assuntos/noticias/enem/para-segundo-dia-exame-registra-maior-indice-de-participacao-da-historia-729')
contentAus2019 = urlAus2019.content

siteAus2019 = BeautifulSoup(contentAus2019, 'html.parser')
tbodyAus2019 = siteAus2019.find('tbody')

urlInsc2020 = requests.get('https://www.gov.br/inep/pt-br/assuntos/noticias/enem/58-milhoes-estao-inscritos-para-fazer-o-enem-2020')
contentInsc2020 = urlInsc2020.content

siteInsc2020 = BeautifulSoup(contentInsc2020, 'html.parser')
tbodyInsc2020 = siteInsc2020.find('tbody')

urlInsc2021 = requests.get('https://www.gov.br/inep/pt-br/assuntos/noticias/enem/divulgados-os-numeros-de-inscritos-no-enem-2021-por-uf')
contentInsc2021 = urlInsc2021.content

siteInsc2021 = BeautifulSoup(contentInsc2021, 'html.parser')
tbodyInsc2021 = siteInsc2021.find('tbody')


# Coleta os dados das tabelas do site
insc2019=[]
aus2019=[]
insc2020=[]
insc2021=[]

for x in tbodyInsc2019.findAll("tr"):
    cells = x.findAll('td')
    if len(cells)==3:
        insc2019.append(cells[1].find('span'))

for x in tbodyAus2019.findAll("tr"):
    cells = x.findAll('td')
    if len(cells)==6:
        aus2019.append(cells[4].find('span'))

for x in tbodyInsc2020.findAll("tr"):
    cells = x.findAll('td')
    if len(cells)==4:
        insc2020.append(cells[3].find(text=True))

for x in tbodyInsc2021.findAll("tr"):
    cells = x.findAll('td')
    if len(cells)==5:
        insc2021.append(cells[4].find('p'))

# Trata os dados recebidos
In2019 = insc2019[26].text
In2019 = float(In2019)*1000
Au2019 = aus2019[26].text
Au2019 = float(Au2019)*1000
In2020 = insc2020[25]
In2020 = float(In2020)*1000
Au2020 = 512460 + 21754
In2021 = insc2021[25].text
In2021 = float(In2021)*1000
Au2021 = 117774 + 10043

# Cria a planilha com os dados
i1=['Ano','Inscrições', 'Ausentes']
i2=['Enem 2019', In2019, Au2019]
i3=['Enem 2020', In2020, Au2020]
i4=['Enem 2021', In2021, Au2021]

planilha = openpyxl.Workbook()
page =planilha['Sheet']
page.title = 'inscricoes'

page.append(i1)
page.append(i2)
page.append(i3)
page.append(i4)

planilha.save("Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/Balanços_Enem-SP.xlsx")

# Lê a planilha
data = pd.read_excel('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/Balanços_Enem-SP.xlsx')

ano = data['Ano']
insc = data['Inscrições']
aus = data['Ausentes']

# Gráfico de inscritos
plt.title("Inscrições no Enem em Sp 2019 a 2021")
plt.ylabel("Inscrições")
plt.xlabel("Anos")
plt.plot(ano, insc, marker='o', color='green')
plt.grid()
plt.savefig('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/Inscriçõs no Enem em Sp 2019-2021')

# Gráfico de ausentes
plt.title("Ausência no Enem em Sp 2019 a 2021")
plt.ylabel("Ausêntes x Inscritos")
plt.xlabel("Anos")
plt.plot(ano, aus, marker='o', color='maroon')
plt.grid()
plt.savefig('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/Ausentes no Enem em Sp 2019-2021')

# Ajusta os meses
def ajustaMes(mes):
    meses = ['jan', 'fev', 'mar', 'abr', 'mai', 'jun', 'jul', 'ago', 'set', 'out', 'nov', 'dez']
    return meses[int(mes)-1]

# Verifica se o valor é do tipo inteiro
def validaInt(str):
  try:
    int(str)
    return True
  except ValueError:
    return False

# Estiliza os gráficos de barras horizontais
def styleBarh(titulo, legendaX, legendaY):
  plt.title(titulo, fontsize=22)
  plt.xlabel(legendaX, fontsize=16)
  plt.ylabel(legendaY, fontsize=16)
  plt.gca().invert_yaxis()
  plt.xticks(fontsize=14)
  plt.yticks(fontsize=14)
  plt.tight_layout()

# Adiciona legenda nas barras horizontais dos gráficos 
def addLegBarh(grafico, soma):
    for reta in grafico:
      altura = reta.get_height()
      largura = reta.get_width()
      plt.text(largura+soma, reta.get_y()+0.6, (str(int(largura))), fontsize=12)

# Link e nome do  CSV
linkSeade = 'https://www.seade.gov.br/wp-content/uploads/coronavirus-files/Dados-covid-19-estado.csv'
planSeade = 'Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/dadosCovidEstadoSP.csv'

# Remove o CSV caso exista
try:
  os.remove(planSeade)
except OSError as e:
  print(f"Erro: {e.strerror}")

# Baixa o CSV 
wget.download(linkSeade, planSeade, False)

# Lê o arquivo e extrai as colunas
tabela = pd.read_csv(planSeade, delimiter=";", encoding='ISO-8859-1')
tabela['Óbitos por dia'].fillna(0, inplace=True)
data = tabela['Data']
casosDiarios = tabela['Casos por dia']
obitosDiarios = tabela['Óbitos por dia']

# Transforma a data diária em mensal
meses = []
for i in range(len(data)):
  [dia, mes, ano] = data[i].split('/')
  if validaInt(mes):
    mes = ajustaMes(mes)
  dado = f'{mes} {ano}'
  if dado not in meses:
    meses.append(dado)

# Coleta os casos de covid mensais
casosMensal = []
qtdCasos = 0
for i in range(len(meses)):
  for j in range(len(data)):
    [dia, mes, ano] = data[j].split('/')
    if validaInt(mes):
      mes = ajustaMes(mes)
    dado = f'{mes} {ano}'
    if dado == meses[i]:
      qtdCasos += casosDiarios[j]
  casosMensal.append(qtdCasos)
  qtdCasos = 0

# Coleta os óbitos de covid mensais
obitosMensal = []
qtdObitos = 0
for i in range(len(meses)):
  for j in range(len(data)):
    [dia, mes, ano] = data[j].split('/')
    if validaInt(mes):
      mes = ajustaMes(mes)
    dado = f'{mes} {ano}'
    if dado == meses[i]:
      qtdObitos += obitosDiarios[j]
  if math.isnan(qtdObitos):
    qtdObitos = 0
  obitosMensal.append(qtdObitos)
  qtdObitos = 0

largura = 17
altura = 11

# Gráfico dos casos
plt.figure(figsize=(largura, altura))
Casos = plt.barh(meses, casosMensal, ec="k", alpha=.6, color="#8a4af3")
styleBarh('Casos por mês - 2020 a 2022', 'Quantidade de casos', 'Data')
addLegBarh(Casos, 900)
plt.savefig('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/casosMensal.png', format='png', dpi=300)

# Gráfico dos óbitos
plt.figure(figsize=(largura, altura))
Obitos = plt.barh(meses, obitosMensal, ec="k", alpha=.6, color="#8a4af3")
styleBarh('Óbitos por mês - 2020 a 2022', 'Quantidade de óbitos', 'Data')
addLegBarh(Obitos, 65)
plt.savefig('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/obitosMensal.png', format='png', dpi=300)

# Cria o PDF
pdf = FPDF('P', 'mm', 'A4')

# Tamanho dos gráficos no PDF
largura = 210
altura = 290

# Cria a capa do relatório
pdf.add_page()
pdf.set_font('Arial', 'B', 18)
pdf.set_text_color(0, 0, 0)
titulo = 'O impacto da COVID19 no ingresso ao ensino superior\nno estado de São Paulo'
pdf.multi_cell(0, 8, titulo, align='C', ln=1)

pdf.set_font('Arial', 'B', 14)
pdf.set_text_color(0, 0, 0)
pdf.cell(0, 210, '', ln=1)
analise = 'Analise feita por:\nJuan Souza'
pdf.multi_cell(0, 8, analise, align='C', ln=1)

# Cria a segunda página
pdf.add_page()
pdf.set_font('Arial', 'B', 16)
pdf.set_text_color(0, 0, 0)
pdf.cell(0, 15, 'Dados da COVID19 em SP', align='C', ln=1)

# Insere os gráficos dos óbitos e casos mensais
pdf.image('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/casosMensal.png', 5, 33, largura-10)
pdf.image('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/obitosMensal.png', 5, 163, largura-10)

# Cria a terceira página
pdf.add_page()
pdf.set_font('Arial', '', 12)
pdf.set_text_color(0, 0, 0)
pdf.cell(0, 20, '', ln=1)

COVID = '   A pandemia de COVID19 atingiu e impactou muitas vidas no mundo todo, direta e indiretamente, e isso não é novidade. No estado de São Paulo não foi diferente, como mostram os gráficos feitos a partir dos dados divulgados pelo governo do estado no site https://www.seade.gov.br/coronavirus/, e em decorrência da crise sanitária tivemos vários problemas na saúde, visto que além de matar a COVID fez com que todos os hospitais ficassem lotados, mas também gerou problemas na economia, os fantasmas da inflação e a fome voltaram a assolar o Brasil e o estado, e impactou negativamente a educação, tivemos varios períodos sem aula e nenhuma política pública eficiênte quanto as aulas EAD.\n\n   Ainda na educação, mesmo após a implementação do ensino a distância a educação não voltou a ser como era antes, e isso impactou e impacta ainda hoje os alunos, principalmente os de escolas públicas, que não tiveram um bom aproveitamento do conteúdo e desde então se sentem inseguros. O impacto no futuro desses jovens pode ser bem observado no Exame Nacional do Ensino Médio (Enem), que é o grande divisor de águas para esses alunos.'
pdf.multi_cell(0, 8, COVID, align='L')

# Cria a quarta página
pdf.add_page()
pdf.set_font('Arial', 'B', 16)
pdf.set_text_color(0, 0, 0)
pdf.cell(0, 15, 'Dados do Enem em SP', align='C', ln=1)

pdf.set_font('Arial', '', 12)

# Insere os gráficos dos inscritos e ausentes
pdf.image('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/Inscriçõs no Enem em Sp 2019-2021.png', 5, 33, (largura/2)-10)
pdf.image('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/Ausentes no Enem em Sp 2019-2021.png', (largura/2)+5, 33, (largura/2)-10)

pdf.cell(0, 100, '', ln=1)
Enem = '   O Enem é o principal objetivo de todo estudante que deseja entrar em uma universidade pública, e além de ser muito importante é normalmente bastante disputado, mas como consequência do impacto da pandemia de COVID19 na educação do estado de São Paulo nós podemos observar que após o pico que a pandemia teve no início e meio do ano de 2020 tivemos um aumento expressivo na ausência dos inscritos na prova do Enem no mesmo ano e em 2021 não só uma grande quantidade de faltas como também uma queda considerável na quantidades de inscritos, como mostram os gráficos acima com informações retiradas tanto do site https://www.gov.br/ quanto de documentos disponibilizados pelo INEP.\n\n   A relevância do Enem está para além de ser o "ingresso" para a universidade, assim como a educação em geral ele é a principal forma de diminuir as desigualdades mas com a crise sanitária o que aconteceu foi um agravamento dessa desigualdade tendo em vista que a população carente não foi só a que mais sofreu com casos da doença, mas também foram as que tiveram o pior projeto de ensino a distância, quando havia um, fazendo assim que a população mais marginalizada fosse quase que excluida da prova.'
pdf.multi_cell(0, 8, Enem, align='L', ln=1)

# Cria a quarta página
pdf.add_page()
pdf.set_font('Arial', 'B', 16)
pdf.set_text_color(0, 0, 0)
pdf.cell(0, 15, 'Resultado', align='C', ln=1)
pdf.cell(0, 20, '', ln=1)

pdf.set_font('Arial', '', 12)
Resultado = '   A pandemia de COVID19 agravou as desigualdades sociais pois causou uma precarização na educação que em decorrência a necessidade de distânciamento social teve que migrar para o digital, mesmo sem a maioria dos estudantes terem a estrutura necessária, e deixou os jovens mais ansiosos e inseguros com relação ao Enem e seu ingresso nas universidades gerando uma diminuição nos números de inscritos e aumento nos números de ausentes nas provas, impossibilitando a entrada desses estudantes nas instituições de ensino superior.'
pdf.multi_cell(0, 8, Resultado, align='L')

# Cria a quinta página
pdf.add_page()
pdf.set_font('Arial', 'B', 16)
pdf.set_text_color(0, 0, 0)
pdf.cell(0, 15, 'Referências Bibliográficas', align='C', ln=1)
pdf.cell(0, 20, '', ln=1)

pdf.set_font('Arial', 'B', 9)

Ref = 'Dados da COVID19 em São Paulo : https://www.seade.gov.br/coronavirus ; \nInscritos no Enem 2019 : https://www.gov.br/inep/pt-br/assuntos/noticias/enem/mais-de-5-milhoes-de-participantes-estao-confirmados-para-a-edicao-do-exame-em-2019 ; \nAusentes no Enem 2019 : https://www.gov.br/inep/pt-br/assuntos/noticias/enem/para-segundo-dia-exame-registra-maior-indice-de-participacao-da-historia-729 ; \nInscritos no Enem 2020: https://www.gov.br/inep/pt-br/assuntos/noticias/enem/58-milhoes-estao-inscritos-para-fazer-o-enem-2020 ; \nAusentes no enem 2020 : https://download.inep.gov.br/educacao_basica/enem/downloads/2020/Enem_Digital_2020_Balanco_Aplicacao_Segundo_Domingo.pdf , https://download.inep.gov.br/educacao_basica/enem/downloads/2020/Enem_Impresso_2020_Balanco_Aplicacao_Segundo_Domingo.pdf ; \nInscritos no Enem 2021 : https://www.gov.br/inep/pt-br/assuntos/noticias/enem/divulgados-os-numeros-de-inscritos-no-enem-2021-por-uf ; \nAusentes no Enem 2021 : https://download.inep.gov.br/enem/outros_documentos/balanco_aplicacao_enem_2021_28112021.pdf .'
pdf.multi_cell(0, 8, Ref, align='L')

# Salva o PDF finalizado
pdf.output('Projeto_Python-O_impacto_da_COVID19_no_ingresso_ao_ensino_superior/O impacto da COVID19 no ingresso ao ensino superior.pdf')

